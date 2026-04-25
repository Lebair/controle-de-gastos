import warnings
warnings.filterwarnings("ignore")

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import json
from pathlib import Path
from openpyxl import load_workbook

# ─── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Controle de Gastos 2026",
    page_icon="💸",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─── Theme palettes ───────────────────────────────────────────────────────────
THEME = {
    False: {  # light
        "bg_page":    "#F1F5F9",
        "bg_card":    "#ffffff",
        "input_bg":   "#ffffff",
        "text_main":  "#0F2C4F",
        "text_muted": "#64748B",
        "border":     "#E2E8F0",
        "accent":     "#1A3F6B",
        "grid":       "#F1F5F9",
        "chart_font": "#334155",
        "chart_bg":   "#ffffff",
    },
    True: {   # dark
        "bg_page":    "#0D1117",
        "bg_card":    "#161B22",
        "input_bg":   "#21262D",
        "text_main":  "#E6EDF3",
        "text_muted": "#8B949E",
        "border":     "#30363D",
        "accent":     "#58A6FF",
        "grid":       "#21262D",
        "chart_font": "#C9D1D9",
        "chart_bg":   "#161B22",
    },
}

# ─── CSS injection ────────────────────────────────────────────────────────────
def inject_css(dark: bool) -> None:
    t = THEME[dark]
    st.markdown(f"""
<style>
/* ── Hide sidebar ── */
[data-testid="stSidebar"],
[data-testid="collapsedControl"] {{ display: none !important; }}

/* ── App-level backgrounds ── */
body,
.stApp,
[data-testid="stAppViewContainer"],
[data-testid="stMain"],
[data-testid="stHeader"],
section.main {{
    background-color: {t['bg_page']} !important;
}}

/* ── Block container ── */
.main .block-container {{
    zoom: 1.08;
    padding-top: 0.5rem !important;
    padding-bottom: 1.2rem !important;
    max-width: 100% !important;
    background-color: {t['bg_page']} !important;
}}

/* ── Scrollable containers (st.container with height) ── */
[data-testid="stVerticalBlockBorderWrapper"] {{
    background-color: transparent !important;
}}

/* ── Global text ── */
[data-testid="stMarkdownContainer"] p,
[data-testid="stMarkdownContainer"] li,
[data-testid="stMarkdownContainer"] span,
[data-testid="stCaptionContainer"] p,
[data-testid="stCaptionContainer"] span,
.stCaption, h1, h2, h3, h4, h5, h6 {{
    color: {t['text_main']} !important;
}}

/* ── Navbar ── */
.navbar {{
    background: linear-gradient(90deg, #0D1B2A 0%, #1E4D8C 100%);
    border-radius: 14px;
    padding: 0 22px;
    height: 66px;
    display: flex;
    align-items: center;
    gap: 14px;
    box-shadow: 0 4px 20px rgba(0,0,0,0.30);
    margin-bottom: 0;
}}
.nav-icon  {{ font-size: 1.7rem; }}
.nav-title {{ color: #fff !important; font-size: 1.18rem; font-weight: 700; line-height: 1.2; }}
.nav-sub   {{ color: rgba(255,255,255,.5) !important; font-size: 0.77rem; }}

/* ── Hamburger / popover button ── */
[data-testid="stPopover"] > button,
[data-testid="stPopover"] > button:focus {{
    background: rgba(255,255,255,0.11) !important;
    border: 1px solid rgba(255,255,255,0.26) !important;
    color: #fff !important;
    border-radius: 10px !important;
    font-size: 1rem !important;
    height: 66px !important;
    width: 100% !important;
    font-weight: 600 !important;
    letter-spacing: 0.02em;
    transition: background 0.15s;
}}
[data-testid="stPopover"] > button:hover {{
    background: rgba(255,255,255,0.20) !important;
}}

/* ── Popover body ── */
[data-testid="stPopoverBody"] {{
    min-width: 360px;
    padding: 18px 22px 14px;
    background-color: {t['bg_card']} !important;
}}
[data-testid="stPopoverBody"] p,
[data-testid="stPopoverBody"] label,
[data-testid="stPopoverBody"] span,
[data-testid="stPopoverBody"] [data-testid="stMarkdownContainer"] p {{
    color: {t['text_main']} !important;
}}

/* ── Metric cards ── */
[data-testid="metric-container"] {{
    background: {t['bg_card']} !important;
    border: 1px solid {t['border']} !important;
    border-radius: 14px;
    padding: 18px 20px 14px !important;
    box-shadow: 0 2px 14px rgba(0,0,0,0.12);
}}
[data-testid="stMetricLabel"] p {{
    font-size: 0.76rem !important;
    color: {t['text_muted']} !important;
    font-weight: 700 !important;
    letter-spacing: 0.05em;
    text-transform: uppercase;
}}
[data-testid="stMetricValue"] {{
    font-size: 1.52rem !important;
    color: {t['text_main']} !important;
    font-weight: 800 !important;
}}

/* ── Section headers ── */
.section-header {{
    background: {t['bg_card']};
    border-radius: 12px;
    padding: 12px 20px;
    margin-bottom: 0.5rem;
    border-left: 4px solid {t['accent']};
    box-shadow: 0 1px 6px rgba(0,0,0,0.08);
    font-size: 1rem;
    font-weight: 700;
    color: {t['text_main']} !important;
}}

/* ── Chart containers ── */
[data-testid="stPlotlyChart"] {{
    background: {t['bg_card']};
    border-radius: 14px;
    padding: 4px;
    box-shadow: 0 2px 12px rgba(0,0,0,0.10);
}}

/* ── Dataframe ── */
[data-testid="stDataFrame"] {{
    border-radius: 12px;
    overflow: hidden;
    box-shadow: 0 2px 12px rgba(0,0,0,0.10);
}}

/* ── Progress bars ── */
.stProgress > div > div > div {{
    background: linear-gradient(90deg, #1A5276, #2ECC71);
    border-radius: 4px;
}}

/* ── Expander ── */
[data-testid="stExpander"] {{
    background: {t['bg_card']} !important;
    border-radius: 12px;
    border: 1px solid {t['border']} !important;
    box-shadow: 0 1px 6px rgba(0,0,0,0.06);
}}
[data-testid="stExpanderDetails"] {{
    background: {t['bg_card']} !important;
}}
[data-testid="stExpander"] summary,
[data-testid="stExpander"] summary p,
[data-testid="stExpander"] summary span {{
    color: {t['text_main']} !important;
}}

/* ── Text input ── */
[data-testid="stTextInput"] input,
[data-baseweb="base-input"] input,
[data-baseweb="input"] input {{
    background-color: {t['input_bg']} !important;
    color: {t['text_main']} !important;
    border-color: {t['border']} !important;
}}
[data-baseweb="input"] {{
    background-color: {t['input_bg']} !important;
    border-color: {t['border']} !important;
}}

/* ── Number input ── */
[data-testid="stNumberInput"] input {{
    background-color: {t['input_bg']} !important;
    color: {t['text_main']} !important;
    border-color: {t['border']} !important;
}}
[data-testid="stNumberInput"] button {{
    background-color: {t['input_bg']} !important;
    color: {t['text_main']} !important;
    border-color: {t['border']} !important;
}}

/* ── Selectbox (inclui dentro do popover portal) ── */
[data-baseweb="select"] {{
    background-color: {t['input_bg']} !important;
}}
[data-baseweb="select"] > div,
[data-baseweb="select"] > div:first-child {{
    background-color: {t['input_bg']} !important;
    border-color: {t['border']} !important;
    color: {t['text_main']} !important;
}}
[data-baseweb="select"] svg {{
    fill: {t['text_muted']} !important;
}}
[data-baseweb="select"] span,
[data-baseweb="select"] div {{
    color: {t['text_main']} !important;
    background-color: {t['input_bg']} !important;
}}
/* Dropdown menu list */
[data-baseweb="menu"],
[data-baseweb="popover"] [data-baseweb="menu"] {{
    background-color: {t['bg_card']} !important;
    border-color: {t['border']} !important;
}}
[data-baseweb="option"],
[role="option"] {{
    background-color: {t['bg_card']} !important;
    color: {t['text_main']} !important;
}}
[data-baseweb="option"]:hover,
[role="option"]:hover {{
    background-color: {t['input_bg']} !important;
}}
/* Popover portal container */
[data-baseweb="popover"] {{
    background-color: {t['bg_card']} !important;
}}
[data-baseweb="popover"] * {{
    border-color: {t['border']} !important;
}}

/* ── Toggle ── */
[data-testid="stToggle"] p,
[data-testid="stToggle"] label,
[data-testid="stToggle"] span {{
    color: {t['text_main']} !important;
}}

/* ── Alert / info / warning ── */
[data-testid="stAlert"] {{
    background-color: {t['bg_card']} !important;
    border-color: {t['border']} !important;
}}
[data-testid="stAlert"] p,
[data-testid="stAlert"] span {{
    color: {t['text_main']} !important;
}}

/* ── Generic buttons ── */
[data-testid="stBaseButton-secondary"] {{
    background-color: {t['input_bg']} !important;
    border-color: {t['border']} !important;
    color: {t['text_main']} !important;
}}

/* ── Divider ── */
hr {{
    border-color: {t['border']} !important;
}}

/* ── Scrollable container do top chart: apenas scroll vertical, sem scroll interno do plotly ── */
[data-testid="stVerticalBlockBorderWrapper"] {{
    overflow-x: hidden !important;
}}
[data-testid="stVerticalBlockBorderWrapper"] [data-testid="stPlotlyChart"],
[data-testid="stVerticalBlockBorderWrapper"] [data-testid="stPlotlyChart"] > div {{
    overflow: hidden !important;
}}

/* ── Scrollbars (webkit — Chrome/Edge) ── */
::-webkit-scrollbar {{
    width: 7px;
    height: 7px;
}}
::-webkit-scrollbar-track {{
    background: {t['bg_page']};
    border-radius: 8px;
}}
::-webkit-scrollbar-thumb {{
    background: {t['border']};
    border-radius: 8px;
    border: 2px solid {t['bg_page']};
}}
::-webkit-scrollbar-thumb:hover {{
    background: {t['text_muted']};
}}
::-webkit-scrollbar-corner {{
    background: {t['bg_page']};
}}

/* ── Widget labels ── */
[data-testid="stWidgetLabel"] p,
[data-testid="stWidgetLabel"] label,
[data-testid="stWidgetLabel"] span {{
    color: {t['text_main']} !important;
}}

/* ── Input placeholder ── */
input::placeholder {{
    color: {t['text_muted']} !important;
    opacity: 0.8;
}}

/* ── Success / info / warning boxes ── */
[data-testid="stNotification"],
[data-testid="stStatusWidget"] {{
    background-color: {t['bg_card']} !important;
    border-color: {t['border']} !important;
}}
[data-testid="stNotification"] p,
[data-testid="stNotification"] span {{
    color: {t['text_main']} !important;
}}
</style>
""", unsafe_allow_html=True)

# ─── Paths & constants ────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent
EXCEL_PATH  = BASE_DIR / "Planilha de controle de gastos 2026.xlsx"
CONFIG_PATH = BASE_DIR / "config.json"

MESES_ORDER = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
MESES_NUM_TO_ABREV = {i + 1: a for i, a in enumerate(MESES_ORDER)}
MESES_FULL = {
    1:"Janeiro", 2:"Fevereiro", 3:"Março",    4:"Abril",
    5:"Maio",    6:"Junho",     7:"Julho",     8:"Agosto",
    9:"Setembro",10:"Outubro",  11:"Novembro", 12:"Dezembro",
}
MESES_FULL_TO_NUM = {v: k for k, v in MESES_FULL.items()}

# Rows that are not real expenses (carry-forward balances, inter-account transfers)
EXCLUDE_KEYWORDS = ["SALDO ANTERIOR", "SALDO ANT"]

CHART_COLORS = [
    "#1A5276","#2ECC71","#E67E22","#9B59B6","#E74C3C",
    "#1ABC9C","#F39C12","#2980B9","#27AE60","#8E44AD",
]

# ─── Helpers ──────────────────────────────────────────────────────────────────
def brl(value: float) -> str:
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def clean_str(s, default: str) -> str:
    return s.strip() if isinstance(s, str) and s.strip() else default

def is_excluded(nome_banco: str | None) -> bool:
    if not nome_banco:
        return False
    nb = nome_banco.upper().strip()
    return any(kw in nb for kw in EXCLUDE_KEYWORDS)


# ─── Data loading ─────────────────────────────────────────────────────────────
@st.cache_data(ttl=300, show_spinner="Carregando planilha...")
def load_data() -> pd.DataFrame:
    if not EXCEL_PATH.exists():
        st.error(f"Planilha não encontrada:\n`{EXCEL_PATH}`"); st.stop()

    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb["Tabela"]

    rows, consecutive_empty = [], 0
    for row_vals in ws.iter_rows(min_row=4, max_col=6, values_only=True):
        if all(v is None or v == "" for v in row_vals):
            consecutive_empty += 1
            if consecutive_empty >= 10:
                break
        else:
            consecutive_empty = 0
            # Skip carry-forward balance rows
            if not is_excluded(str(row_vals[1]) if row_vals[1] else None):
                rows.append(row_vals)
    wb.close()

    if not rows:
        st.error("A aba 'Tabela' não tem dados ou está em formato inesperado."); st.stop()

    df = pd.DataFrame(rows, columns=["Data","Nome banco","Nome","Categoria","Condição","Valor"])
    df["Data"]  = pd.to_datetime(df["Data"], errors="coerce")
    df["Valor"] = pd.to_numeric(df["Valor"], errors="coerce")
    df = df.dropna(subset=["Data","Valor"])
    df = df[df["Valor"] > 0]

    df["Nome"]      = df.apply(lambda r: clean_str(r["Nome"],"") or clean_str(r["Nome banco"],"Desconhecido"), axis=1)
    df["Categoria"] = df["Categoria"].apply(lambda x: clean_str(x, "Sem categoria"))
    df["Condição"]  = df["Condição"].apply(lambda x: clean_str(x, "Não informado"))
    df["Mês_num"]   = df["Data"].dt.month
    df["Mês_abrev"] = df["Mês_num"].map(MESES_NUM_TO_ABREV)
    df["Mês_nome"]  = df["Mês_num"].map(MESES_FULL)

    return df.sort_values("Data").reset_index(drop=True)


def compute_monthly_totals(df: pd.DataFrame) -> dict[str, float]:
    """Calcula totais mensais diretamente da aba Tabela."""
    return (
        df.groupby("Mês_nome")["Valor"]
        .sum()
        .to_dict()
    )


# ─── Config (metas) ───────────────────────────────────────────────────────────
def load_config() -> dict:
    if CONFIG_PATH.exists():
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"metas": {}}

def save_config(config: dict) -> None:
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)


# ─── Main ─────────────────────────────────────────────────────────────────────
def main() -> None:
    # ── Restaurar dark_mode da URL antes de injetar CSS ──────────────────────
    qp = st.query_params
    if "dark" in qp and "dark_mode" not in st.session_state:
        st.session_state["dark_mode"] = (qp["dark"] == "1")
    elif "dark_mode" not in st.session_state:
        st.session_state["dark_mode"] = True  # padrão: modo escuro

    dark = st.session_state.get("dark_mode", True)
    inject_css(dark)

    t = THEME[dark]  # shortcut for chart colors

    df             = load_data()
    monthly_totals = compute_monthly_totals(df)
    config         = load_config()

    # ── Build filter options ──────────────────────────────────────────────────
    # Period: "Todos os meses" + "Mês / Ano" per month in data
    years_in_data = sorted(df["Data"].dt.year.unique())
    period_opts   = ["Todos os meses"]
    for yr in years_in_data:
        for m in sorted(df[df["Data"].dt.year == yr]["Mês_num"].unique()):
            period_opts.append(f"{MESES_FULL[m]} / {yr}")

    cat_opts  = ["Todas as categorias"] + sorted(df["Categoria"].unique())
    cond_opts = ["Todas as condições"]  + sorted(df["Condição"].unique())

    # ── Restaurar filtros da URL (só na primeira execução após F5) ────────────
    if "period" in qp and qp["period"] in period_opts and "filter_period" not in st.session_state:
        st.session_state["filter_period"] = qp["period"]
    if "cat" in qp and qp["cat"] in cat_opts and "filter_cat" not in st.session_state:
        st.session_state["filter_cat"] = qp["cat"]
    if "cond" in qp and qp["cond"] in cond_opts and "filter_cond" not in st.session_state:
        st.session_state["filter_cond"] = qp["cond"]

    # ── Navbar + hamburger popover ────────────────────────────────────────────
    col_nav, col_menu = st.columns([11, 1])

    with col_nav:
        periodo_nav = st.session_state.get("filter_period", "Todos os meses")
        st.markdown(f"""
        <div class="navbar">
            <span class="nav-icon">💸</span>
            <div>
                <div class="nav-title">Controle de Gastos</div>
                <div class="nav-sub">{periodo_nav}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    with col_menu:
        with st.popover("☰ Filtros", use_container_width=True):
            st.markdown("**📅 Período**")
            st.selectbox("Período", period_opts,
                         key="filter_period", label_visibility="collapsed")

            st.markdown("**🏷️ Categoria**")
            st.selectbox("Categoria", cat_opts,
                         key="filter_cat", label_visibility="collapsed")

            st.markdown("**💳 Condição**")
            st.selectbox("Condição", cond_opts,
                         key="filter_cond", label_visibility="collapsed")

            st.divider()
            st.toggle("🌙 Modo escuro", key="dark_mode")
            if st.button("↺ Recarregar dados", use_container_width=True):
                st.cache_data.clear(); st.rerun()

    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

    # ── Read filters ──────────────────────────────────────────────────────────
    selected_period = st.session_state.get("filter_period", "Todos os meses")
    sel_cat         = st.session_state.get("filter_cat",  "Todas as categorias")
    sel_cond        = st.session_state.get("filter_cond", "Todas as condições")

    # ── Filtering ─────────────────────────────────────────────────────────────
    mask = pd.Series(True, index=df.index)
    if sel_cat  != "Todas as categorias":
        mask &= df["Categoria"] == sel_cat
    if sel_cond != "Todas as condições":
        mask &= df["Condição"]  == sel_cond
    df_base = df[mask]

    month_num = None
    year_num  = None
    if selected_period != "Todos os meses":
        parts     = selected_period.split(" / ")
        month_num = MESES_FULL_TO_NUM[parts[0]]
        year_num  = int(parts[1])
        df_full   = df_base[
            (df_base["Mês_num"] == month_num) &
            (df_base["Data"].dt.year == year_num)
        ].copy()
    else:
        df_full = df_base.copy()

    if df_full.empty:
        st.warning("⚠️ Nenhum registro com os filtros aplicados."); st.stop()

    # ── Totals (use pre-computed Valores sheet when no category filter) ────────
    cat_filter_active = sel_cat != "Todas as categorias"
    if not cat_filter_active and month_num:
        total   = df_full["Valor"].sum()
        n_meses = 1
    elif not cat_filter_active:
        total   = df_full["Valor"].sum()
        n_meses = df_full["Mês_num"].nunique()
    else:
        total   = df_full["Valor"].sum()
        n_meses = max(df_full["Mês_num"].nunique(), 1)

    maior        = df_full["Valor"].max()
    n_trans      = len(df_full)
    media_mensal = total / n_meses if n_meses > 0 else total

    # ── Cards ─────────────────────────────────────────────────────────────────
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("💰 Total Gasto",  brl(total))
    c2.metric("📅 Média Mensal", brl(media_mensal))
    c3.metric("🔝 Maior Gasto",  brl(maior))
    c4.metric("📋 Transações",   str(n_trans))

    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

    # ── Evolução Mensal ────────────────────────────────────────────────────────
    st.markdown("<div class='section-header'>📈 Evolução Mensal</div>", unsafe_allow_html=True)

    monthly_rows = []
    for mn in sorted(df["Mês_num"].unique()):
        mes_nome  = MESES_FULL[mn]
        mes_abrev = MESES_NUM_TO_ABREV[mn]
        valor     = df_base[df_base["Mês_num"] == mn]["Valor"].sum()
        monthly_rows.append({"Mês_num": mn, "Mês_abrev": mes_abrev, "Mês_nome": mes_nome, "Valor": valor})

    monthly_df = pd.DataFrame(monthly_rows)
    monthly_df["Mês_abrev"] = pd.Categorical(monthly_df["Mês_abrev"], categories=MESES_ORDER, ordered=True)
    monthly_df = monthly_df.sort_values("Mês_abrev")

    bar_colors = [
        "#E67E22" if (month_num and row["Mês_nome"] == MESES_FULL.get(month_num)) else "#1A5276"
        for _, row in monthly_df.iterrows()
    ]
    avg_line = monthly_df["Valor"].mean() if len(monthly_df) > 0 else 0

    fig_bar = go.Figure()
    fig_bar.add_trace(go.Bar(
        x=monthly_df["Mês_abrev"], y=monthly_df["Valor"],
        marker_color=bar_colors,
        hovertemplate="<b>%{x}</b><br>Total: R$ %{y:,.2f}<extra></extra>",
        text=[brl(v).replace("R$ ", "") for v in monthly_df["Valor"]],
        textposition="outside", textfont=dict(size=13, color=t["chart_font"]),
    ))
    if len(monthly_df) > 1:
        fig_bar.add_hline(
            y=avg_line, line_dash="dot", line_color="#E74C3C", line_width=1.5,
            annotation_text=f"Média: {brl(avg_line)}",
            annotation_position="top right",
            annotation_font_color="#E74C3C",
        )
    fig_bar.update_layout(
        margin=dict(t=30, b=10, l=10, r=10),
        plot_bgcolor=t["chart_bg"], paper_bgcolor=t["chart_bg"],
        yaxis=dict(tickformat=",.0f", title="R$", gridcolor=t["grid"],
                   showline=False, color=t["chart_font"]),
        xaxis=dict(showgrid=False, color=t["chart_font"]),
        showlegend=False, height=300,
        font=dict(color=t["chart_font"]),
    )
    st.plotly_chart(fig_bar, use_container_width=True)

    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

    # ── Category donut + Top categories ───────────────────────────────────────
    chart_cat_filter = None
    col_pie, col_top = st.columns([1, 1])

    with col_pie:
        st.markdown("<div class='section-header'>🏷️ Por Categoria</div>", unsafe_allow_html=True)
        cat_totals = (
            df_full.groupby("Categoria")["Valor"].sum()
            .reset_index().sort_values("Valor", ascending=False)
        )
        fig_pie = px.pie(
            cat_totals, values="Valor", names="Categoria", hole=0.45,
            color_discrete_sequence=CHART_COLORS,
        )
        fig_pie.update_traces(
            textinfo="label+percent",
            hovertemplate="<b>%{label}</b><br>R$ %{value:,.2f}<br>%{percent}<extra></extra>",
            textfont=dict(size=14, color=t["chart_font"]),
        )
        fig_pie.update_layout(
            margin=dict(t=10, b=10, l=10, r=10), showlegend=False,
            plot_bgcolor=t["chart_bg"], paper_bgcolor=t["chart_bg"],
            font=dict(size=14, color=t["chart_font"]),
            height=460,
        )
        st.plotly_chart(fig_pie, use_container_width=True)

    with col_top:
        st.markdown("<div class='section-header'>🔝 Maiores Gastos</div>", unsafe_allow_html=True)
        st.caption("💡 Clique em uma barra para filtrar as transações abaixo")
        top_cats = (
            df_full.groupby("Categoria")["Valor"].sum()
            .reset_index().sort_values("Valor", ascending=True)
        )
        n_cats = len(top_cats)
        chart_height_top = max(500, n_cats * 42 + 80)
        fig_top = go.Figure(go.Bar(
            x=top_cats["Valor"], y=top_cats["Categoria"], orientation="h",
            marker_color="#1A5276",
            text=[brl(v) for v in top_cats["Valor"]],
            textposition="outside",
            textfont=dict(size=13, color=t["chart_font"]),
            hovertemplate="<b>%{y}</b><br>R$ %{x:,.2f}<extra></extra>",
        ))
        fig_top.update_layout(
            margin=dict(t=10, b=10, l=10, r=110),
            plot_bgcolor=t["chart_bg"], paper_bgcolor=t["chart_bg"],
            xaxis=dict(showgrid=True, gridcolor=t["grid"], tickformat=",.0f",
                       title="", color=t["chart_font"], tickfont=dict(size=13)),
            yaxis=dict(showgrid=False, title="", color=t["chart_font"], tickfont=dict(size=13)),
            font=dict(size=13, color=t["chart_font"]),
            height=chart_height_top, showlegend=False,
            clickmode="event+select",
        )
        with st.container(height=460, border=False):
            event_top = st.plotly_chart(
                fig_top, use_container_width=True,
                on_select="rerun", key="chart_top",
                selection_mode="points",
            )
        pts = (event_top.selection.points if event_top and event_top.selection else [])
        chart_cat_filter = pts[0].get("y") if pts else None

    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

    # ── Metas por categoria ────────────────────────────────────────────────────
    st.markdown("<div class='section-header'>🎯 Metas por Categoria</div>", unsafe_allow_html=True)

    metas     = config.get("metas", {})
    cat_gasto = df_full.groupby("Categoria")["Valor"].sum().to_dict()

    with st.expander("✏️ Definir / editar metas mensais"):
        all_cats_meta = sorted(df["Categoria"].unique())
        edit_cols     = st.columns(4)
        new_metas: dict[str, float] = {}
        for i, cat in enumerate(all_cats_meta):
            current_val = float(metas.get(cat, 0.0))
            new_val = edit_cols[i % 4].number_input(
                cat, min_value=0.0, value=current_val, step=50.0,
                key=f"meta_{cat}", format="%.2f",
            )
            new_metas[cat] = new_val
        if st.button("💾 Salvar metas", type="primary"):
            config["metas"] = new_metas
            save_config(config)
            st.success("✅ Metas salvas!")

    cats_com_meta = [c for c in sorted(cat_gasto) if float(metas.get(c, 0)) > 0]
    if cats_com_meta:
        prog_cols = st.columns(min(4, len(cats_com_meta)))
        for i, cat in enumerate(cats_com_meta):
            gasto = cat_gasto.get(cat, 0.0)
            meta  = float(metas[cat])
            pct   = gasto / meta
            icon  = "🟢" if pct < 0.75 else ("🟡" if pct < 1.0 else "🔴")
            with prog_cols[i % 4]:
                st.markdown(f"**{icon} {cat}**")
                st.progress(min(pct, 1.0))
                st.caption(f"{brl(gasto)} / {brl(meta)} ({pct*100:.0f}%)")
    else:
        st.info("📌 Defina metas mensais por categoria no painel acima para acompanhar seu orçamento.")

    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

    # ── Tabela detalhada ──────────────────────────────────────────────────────
    st.markdown("<div class='section-header'>📋 Transações Detalhadas</div>", unsafe_allow_html=True)

    if chart_cat_filter:
        col_info, col_clear = st.columns([9, 1])
        col_info.info(f"📊 Categoria selecionada: **{chart_cat_filter}** — clique na barra novamente para remover")
        if col_clear.button("✕ Limpar", key="clear_chart_cat"):
            chart_cat_filter = None

    search = st.text_input("🔍 Buscar transação", placeholder="Nome, categoria...", label_visibility="collapsed")

    tbl = df_full[["Data", "Nome", "Categoria", "Condição", "Valor"]].copy()
    tbl = tbl.sort_values("Data", ascending=False).reset_index(drop=True)

    if chart_cat_filter:
        tbl = tbl[tbl["Categoria"] == chart_cat_filter]

    if search.strip():
        mask_s = (
            tbl["Nome"].str.contains(search, case=False, na=False)
            | tbl["Categoria"].str.contains(search, case=False, na=False)
        )
        tbl = tbl[mask_s]

    disp = tbl.copy()
    disp["Data"]  = disp["Data"].dt.strftime("%d/%m/%Y")
    disp["Valor"] = disp["Valor"].apply(brl)
    disp.index    = range(1, len(disp) + 1)
    disp.index.name = "#"

    table_html = disp.to_html(classes="expense-table", border=0, escape=True)
    st.markdown(f"""
<div style="overflow:auto;max-height:430px;border:1px solid {t['border']};border-radius:12px;margin-bottom:4px">
<style>
.expense-table {{width:100%;border-collapse:collapse;font-size:0.88rem}}
.expense-table th {{background:{t['bg_card']};color:{t['text_main']};padding:8px 12px;text-align:left;border-bottom:2px solid {t['border']};position:sticky;top:0;white-space:nowrap;font-weight:700}}
.expense-table td {{padding:7px 12px;color:{t['text_main']};border-bottom:1px solid {t['border']};background:{t['bg_page']}}}
.expense-table tr:nth-child(even) td {{background:{t['bg_card']}}}
.expense-table tr:hover td {{background:{t['input_bg']}}}
</style>
{table_html}
</div>
""", unsafe_allow_html=True)
    st.caption(f"Exibindo {len(disp):,} de {len(df_full):,} transações")

    # ── Persistir filtros na URL (sobrevive ao F5) ────────────────────────────
    st.query_params["period"] = st.session_state.get("filter_period", "Todos os meses")
    st.query_params["cat"]    = st.session_state.get("filter_cat",    "Todas as categorias")
    st.query_params["cond"]   = st.session_state.get("filter_cond",   "Todas as condições")
    st.query_params["dark"]   = "1" if st.session_state.get("dark_mode", False) else "0"


if __name__ == "__main__":
    main()

